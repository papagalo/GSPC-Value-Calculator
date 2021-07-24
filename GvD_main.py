from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# Create a workbook object
wb = Workbook()

# load existing spreadsheet
wb = load_workbook(filename='Resources/GSPCdateAsText.xlsx', data_only=True)
ws = wb.active

date_range = ws['A2':'A2145']
gspc_close_range = ws['B2':'B2145']
gspc_dc_range = ws['C3':'C2145']
davt_close_range = ws['E2':'E2145']
davt_dc_range = ws['F2':'F2145']

dateList = []
gspcCloseList = []
gspcDailyChangeList = []
davtCloseList = []
davtDailyChangeList = []

# Iterates through the passed dailyChange list, building a string of output to
# display streaks
def findStreaksDailyChange(streakLength, minValue, dates, dailyChange):
	streakDates = ""
	runningDates = ""
	runningStreakCounter = 0
	index = 0

	for val in dailyChange:
		index += 1
		if val >= minValue:
			runningDates += dates[index] + '\n'
			runningStreakCounter += 1
		else:
			if runningStreakCounter >= streakLength:
				streakDates += runningDates + '\n'
			runningDates = ""
			runningStreakCounter = 0
	
	print(streakDates)


# "Load" functions load their respective lists with data from the Excel sheet
def load_dates(worksheet):
	for cell in date_range:
		for x in cell:
			dateList.append(x.value)
	
def load_gspc_close(worksheet):
	for cell in gspc_close_range:
		for x in cell:
			gspcCloseList.append(x.value)

def load_gspc_daily_change(worksheet):
	for cell in gspc_dc_range:
		for x in cell:
			gspcDailyChangeList.append(x.value * 100)

def load_davt_close(worksheet):
	for cell in davt_close_range:
		for x in cell:
			davtCloseList.append(x.value)

def load_davt_daily_change(worksheet):
	for cell in davt_dc_range:
		for x in cell:
			davtDailyChangeList.append(x.value)

load_dates(ws)
load_gspc_daily_change(ws)
#findStreaksDailyChange(2, 1, dateList, gspcDailyChangeList)
