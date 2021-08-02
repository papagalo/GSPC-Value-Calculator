import tkinter as tk
import pdb
import sys
from tkinter import ttk
from tkinter import scrolledtext
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# Create a workbook object
wb = Workbook()

sys.stdout = open('gspcVSdavt results.txt', 'w')

# load existing spreadsheet
wb = load_workbook(filename='GSPCdateAsText.xlsx', data_only=True)
ws = wb.active
root = tk.Tk()
root.title("GvD Calculator")
root.geometry("400x225")
v = tk.IntVar()

# Declare ranges to load lists
date_range = ws['A3':'A2145']
gspc_close_range = ws['B3':'B2145']
gspc_dc_range = ws['C3':'C2145']
davt_close_range = ws['E3':'E2145']
davt_dc_range = ws['F3':'F2145']

# Declare lists
dateList = []
gspcCloseList = []
gspcDailyChangeList = []
davtCloseList = []
davtDailyChangeList = []
differentials = []

# Iterates through the passed dailyChange list, building a string of output to
# display streaks
def findStreakDifferential():
	streakLength = eStreak.get()
	minValue = eMin.get()
	futureDays = eFuture.get()

	minV = float(minValue)
	streak = int(streakLength)
	future_days = int(futureDays)

	streakDateIndexList = calc_Streak(minV, streak, future_days)
	
	write_output(streakDateIndexList, minValue, streakLength, future_days)
	
	sys.stdout.close()
	root.destroy
	root.quit()

# streak_dates should be loaded with the current streak.
# day_count should be the 'Z' value or the # of days into
# 	the future that you want to look
# start is the index from the current iteration through the
# 	differentials list
# this function will append the next Z days to the current
# 	list of dates
def add_future_days(streak_date_indexes, future_day_count, start):
	end = start + future_day_count
	for x in range(start, end):
		streak_date_indexes.append(x)
	return streak_date_indexes

# This function takes the minimum value the user wants the differential to be
# 	and the minimum streak they want, and finally the amount of days past the
# 		streak that they want to see values of
# Calculates the days that meet the criteria, and returns them in a list of
# 	ints that represent the index of the main lists of the program
def calc_Streak(minV, streak, future_day_count):
	streakDateIndexes = []
	runningDateIndexes = []
	runningStreakCounter = 0
	index = 0

	for val in differentials:
		
		if val >= minV:
			# Save the date at the index point in a temp var
			runningDateIndexes.append(index)
			runningStreakCounter += 1
		else:
			# Streak is broken. Save it to streakDates (what gets printed
			# 	later) and reset temp vars
			if runningStreakCounter >= streak:
				# append the flag to signal the end of a streak
				runningDateIndexes.append(-2)
				# add the next future_day_count worth of dates to the main list
				streakDateIndexes.extend(add_future_days(runningDateIndexes, future_day_count, index))
				# append another flag to signal the end of a streak & future days
				streakDateIndexes.append(-1)
			runningDateIndexes = []
			runningStreakCounter = 0

		index += 1

	return streakDateIndexes

# Writes the output in a new window
def write_output(indexList, minValue, streakLength, future_days):
	window = tk.Tk()
	window.title("Results")

	introString = ('Dates where the differential ("^gspc" daily change - ' +
		'daily account" daily change) was higher than ' + 
			minValue + '% for ' + streakLength + ' or more days in a row:\n\n')

	mainString = prettyStringBuilder(indexList, minValue, streakLength, future_days)

	text_area = scrolledtext.ScrolledText(window,
										  wrap = tk.WORD,
										  width = 50,
										  height = 40,
										  font = ("Times New Roman", 12))
	text_area.grid(column = 0, pady = 10, padx = 10)
	text_area.insert(tk.INSERT, introString)
	text_area.insert(tk.INSERT, mainString)

	window.mainloop()

# "Load" functions load their respective lists with data from the Excel sheet
def load_dates(worksheet):
	for cell in date_range:
		for x in cell:
			dateList.append(x.value)
	
def load_gspc_close(worksheet):
	for cell in gspc_close_range:
		for x in cell:
			gspcCloseList.append(x.value)

def load_gspc_daily_change(worksheet):
	for cell in gspc_dc_range:
		for x in cell:
			gspcDailyChangeList.append(x.value * 100)

def load_davt_close(worksheet):
	for cell in davt_close_range:
		for x in cell:
			davtCloseList.append(x.value)

def load_davt_daily_change(worksheet):
	for cell in davt_dc_range:
		for x in cell:
			davtDailyChangeList.append(x.value * 100)

def load_differentials(gspcDailyChange, davtDailyChange):
	for i in range(2143):
		differentials.append(gspcDailyChange[i] - davtDailyChange[i])

# Call all the load functions
def load_everything(worksheet):
	load_dates(ws)
	load_gspc_close(ws)
	load_gspc_daily_change(ws)
	load_davt_close(ws)
	load_davt_daily_change(ws)
	load_differentials(gspcDailyChangeList, davtDailyChangeList)

def prettyPrint(indexList, minValue, streakLength):
	print('Dates where the differential ("^gspc" daily change - ' +
		'daily account" daily change) was \nhigher than ' + 
			minValue + '% for ' + streakLength + ' or more days in a row:\n')

	for i in indexList:
		print(f"{dateList[i]}:\n^GSPC closing value was: {gspcCloseList[i]}" +
			f"\n^GSPC Percent Change was: {gspcDailyChangeList[i]}\n" +
				f"Daily Account closing value was: {davtCloseList[i]}\n" +
					"the Daily Account Percent Change was: " +
					f"{davtDailyChangeList[i]}\nThe differential was: " +
					f"{differentials[i]}\n")

# Creates a main string to display output in the scrolled text box
def prettyStringBuilder(indexList, minValue, streakLength, future_days):
	curr_index = 0
	firstStreakLen = currStreakLength(curr_index, indexList)
	mainString = (f"New {firstStreakLen} day long streak where the " +
		f"differential was at least {minValue}%\n")
	for i in indexList:
		if (i == -1):
			currStreakLen = currStreakLength(curr_index + 1, indexList)
			#currStreakLen = streakLength
			mainString += (f"\n\nNew {currStreakLen} day long streak where the " +
			f"differential was at least {minValue}%\n")
		elif (i == -2):
			mainString += (f"\nThe next {future_days} days had these stats:\n")
		else:
			mainString += (f"{dateList[i]}:\n^GSPC closing value was: {gspcCloseList[i]}" +
				f"\n^GSPC Percent Change was: {gspcDailyChangeList[i]}\n" +
				f"Daily Account closing value was: {davtCloseList[i]}\n" +
				"the Daily Account Percent Change was: " +
				f"{davtDailyChangeList[i]}\nThe differential was: " +
				f"{differentials[i]}\n")
		curr_index += 1
	return mainString


def currStreakLength(index, index_list):
	length = 0

	for i in range(index,len(index_list)):
		if (index_list[i] == -2):
			return length
		else:
			length += 1
	
	return length

# End definitions
# Begin program logic

# Load the arrays with data
load_everything(ws)

# Set up the GUI

# Set up the min daily change label and entry box
minDailyChange_label = tk.Label(root, text = "Minimum Daily Change:")
minDailyChange_label.grid(row = 1, column = 0, padx = 10, pady = 10)

eMin = tk.Entry(root, width = 10, borderwidth = 4)
eMin.grid(row = 1, column = 1, padx = 10, pady = 10, columnspan = 2)

# Set up the min daily change label and entry box
streak_label = tk.Label(root, text = "How many days in a row?")
streak_label.grid(row = 2, column = 0, padx = 10, pady = 10)

eStreak = tk.Entry(root, width = 10, borderwidth = 4)
eStreak.grid(row = 2, column = 1, padx = 10, pady = 10, columnspan = 2)

# Set up the min daily change label and entry box
future_label = tk.Label(root, text = "Display difference in % change over" +
	" next X days:")
future_label.grid(row = 3, column = 0, padx = 10, pady = 10)

eFuture = tk.Entry(root, width = 10, borderwidth = 4)
eFuture.grid(row = 3, column = 1, padx = 10, pady = 10, columnspan = 2)

# Create submit button
submitButton = tk.Button(root, text = "Submit", padx = 40, pady = 20, 
	width = 1, height = 1, command = lambda: findStreakDifferential())
submitButton.grid(row = 4, column = 0)

# Create an exit button that kills the program
exitButton = tk.Button(root, text = "Exit", padx = 40, pady = 20, 
	width = 1, height = 1, command = root.destroy)
exitButton.grid(row = 4, column = 1)


root.mainloop()